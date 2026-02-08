# ====================================================================
# 1. CONFIGURAÇÃO MANDATÓRIA DO AMBIENTE DJANGO
# ====================================================================
# import os
# import django
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ilustracoes.settings')
# django.setup()

# ====================================================================
# 2. IMPORTAÇÕES DE MODELOS
# ====================================================================
from planilha.models import Ilustracao, Ilustrador, Credito, Projeto, Componente

# ====================================================================
# 3. PACOTES ADICIONAIS
# ====================================================================
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


def get_verbose_names(model_class):
    """Retorna uma lista de verbose_names para os campos do modelo."""
    verbose_names = []
    # O atributo '_meta' contém metadados do modelo
    for field in model_class._meta.fields:
        # Cada campo tem um atributo 'verbose_name'
        verbose_names.append(field.verbose_name)
    return verbose_names

def get_choice_options(model_class):
    """
    Retorna um dicionário mapeando o verbose_name do campo
    para a lista de opções (Choices) ou para a classe do campo
    se for uma ForeignKey (para indicar que é uma lista dinâmica).
    """
    options = {}
    for field in model_class._meta.fields:
        # 1. Trata campos com Choices (como Status, Categoria, etc.)
        if field.choices:
            # Pega o 'Display Name' (segundo elemento da tupla)
            options[field.verbose_name] = [choice[1] for choice in field.choices]

        # 2. Trata ForeignKeys (como Ilustrador)
        elif isinstance(field, (Ilustracao._meta.get_field('ilustrador').__class__,)):
            # Usando a classe do campo para identificar o tipo (models.ForeignKey)
            # Retorna o nome do modelo referenciado
            options[field.verbose_name] = field.related_model.__name__

    return options

def get_foreign_key_values(model_name):
    """Retorna uma lista de strings de todos os objetos de um modelo."""
    if model_name == 'Ilustrador':
        # 1. Pega todos os objetos Ilustrador (Ilustrador.objects.all())
        # 2. Ordena por 'nome'
        # 3. Usa list comprehension para aplicar a função str() em cada objeto,
        #    o que implicitamente chama o método __str__(self) do modelo,
        #    retornando o formato desejado (sigla - nome).
        return [str(ilustrador) for ilustrador in Ilustrador.objects.all().order_by('nome')]
    
    # Lógica para outros modelos (mantida inalterada, presumindo que eles retornam 'nome' em __str__)
    elif model_name == 'Projeto':
        return list(Projeto.objects.values_list('nome', flat=True).order_by('nome'))
    elif model_name == 'Componente':
        return list(Componente.objects.values_list('nome', flat=True).order_by('nome'))
    elif model_name == 'Credito':
        # Se 'Credito' também usa o formato __str__, você deve mudar este aqui também:
        return list(Credito.objects.values_list('nome', flat=True).order_by('nome'))

    return []

def reordenar_colunas_para_excel(colunas_originais):
    """
    Reordena a lista de verbose_names do modelo Ilustracao para corresponder 
    à ordem da tabela HTML e exclui os campos de metadados.
    """
    
    # 1. Define a ORDEM DESEJADA conforme o template HTML (usando os verbose_names)
    # Campos que devem aparecer primeiro e na ordem especificada.
    ORDEM_TEMPLATE = [
        'Retranca', 
        'Status', 
        'Categoria', 
        'Localização', 
        'Volume', 
        'Unidade', 
        'Capítulo ou seção', 
        'Página', 
        'Tipo', 
        'Descrição', 
        'Observação editorial e núcleo',
        'Lote',
        'Data de liberação do lote',
        'Data de envio do pedido',
        'Data de recebimento do rafe',
        'Data de retorno do rafe',
        'Data de recebimento da finalizada',
        'Classificação', 
        'Crédito', 
        'Ilustrador resgate', 
        'Ilustrador criação', 
        'Ilustrador ajuste', 
        'Observação da arte',
        'Pagamento',
        'Projeto',
        'Componente',
    ]

    # 2. Define os campos que devem ser EXCLUÍDOS (metadados e chaves)
    CAMPOS_EXCLUIDOS = [
        'ID', 
        'Ativo', 
        'Criado por',
        'Criado em',
        'Atualizado por',
        'Modificado em',
    ]

    # 3. Converte a lista original em um conjunto (set) para busca rápida
    colunas_originais_set = set(colunas_originais)
    
    colunas_reordenadas = []

    # 4. Adiciona as colunas na ORDEM_TEMPLATE
    for coluna_nome in ORDEM_TEMPLATE:
        if coluna_nome in colunas_originais_set and coluna_nome not in CAMPOS_EXCLUIDOS:
            colunas_reordenadas.append(coluna_nome)
            
    # Opcional: Adicionar quaisquer colunas restantes no final (caso alguma nova 
    # coluna seja adicionada no modelo e não esteja no ORDEM_TEMPLATE)
    colunas_ja_adicionadas = set(colunas_reordenadas)
    
    for coluna in colunas_originais:
        if coluna not in colunas_ja_adicionadas and coluna not in CAMPOS_EXCLUIDOS:
            colunas_reordenadas.append(coluna)
            
    return colunas_reordenadas

def create_excel():
    COLUNAS_ILUSTRACAO0 = get_verbose_names(Ilustracao)
    COLUNAS_ILUSTRACAO = reordenar_colunas_para_excel(COLUNAS_ILUSTRACAO0)

    # ====================================================================
    # 4. CONFIGURAÇÃO E GERAÇÃO DO EXCEL
    # ====================================================================

    # 1. Obter todas as opções de dropdown
    field_options = get_choice_options(Ilustracao)
    dynamic_options = {}

    for verbose_name, value in field_options.items():
        if isinstance(value, str): # Se for o nome de um modelo (ex: 'Ilustrador')
            # Busca os valores no banco de dados
            dynamic_options[verbose_name] = get_foreign_key_values(value)

    # 2. Criar o Workbook e as Worksheets
    workbook = openpyxl.Workbook()
    ws_data = workbook.active
    ws_data.title = "Dados de Ilustracao"

    ws_options = workbook.create_sheet(title="Opcoes_Dropdown")
    ws_options.sheet_state = 'hidden' # Oculta a planilha

    # 3. Configurar a planilha de opções (ws_options)
    option_col = 1
    dropdown_references = {}
    all_dropdown_fields = {**field_options, **dynamic_options}

    for verbose_name, options_list in all_dropdown_fields.items():
        
        # Verifica se é uma lista válida para dropdown
        if isinstance(options_list, list) and options_list:
            
            ws_options.cell(row=1, column=option_col, value=verbose_name)
            
            # Preenche a coluna com as opções
            for row_num, option in enumerate(options_list, 2):
                ws_options.cell(row=row_num, column=option_col, value=option)

            # Início da criação do Defined Name:
            # 1. Define o intervalo absoluto (ex: $B$2:$B$11)
            list_range = f'${get_column_letter(option_col)}$2:${get_column_letter(option_col)}${len(options_list) + 1}'
            
            # 2. Define um nome de referência
            dropdown_name = verbose_name.replace(' ', '_').replace('á', 'a').replace('é', 'e').replace('ç', 'c') + "_List"
            
            # 3. CORREÇÃO DE SINTAXE DO OPENPYXL: Usa o objeto DefinedName
            # A sintaxe f"'{ws_options.title}'!{list_range}" garante que a referência 
            # seja interpretada corretamente pelo Excel (ex: 'Opcoes_Dropdown'!$B$2:$B$11)
            defined_name = DefinedName(
                name=dropdown_name, 
                attr_text=f'\'{ws_options.title}\'!{list_range}' 
            )
            workbook.defined_names.add(defined_name)
            
            dropdown_references[verbose_name] = dropdown_name
            
            option_col += 1

    # 4. Configurar a planilha principal (ws_data) - Cabeçalhos e Validação
    ws_data.append(COLUNAS_ILUSTRACAO)
    ws_data.freeze_panes = 'A2'

    START_ROW = 2
    END_ROW = 500

    for col_idx, verbose_name in enumerate(COLUNAS_ILUSTRACAO, 1):
        col_letter = get_column_letter(col_idx)

        # 5. Aplica Validação de Dados (Dropdowns)
        if verbose_name in dropdown_references:
            formula = f'={dropdown_references[verbose_name]}'
            
            dv = DataValidation(
                type="list", 
                formula1=formula, # Referencia o nome definido (ex: '=Categoria_List')
                allow_blank=True,
                showErrorMessage=True,
                errorTitle='Valor Inválido',
                error='O valor inserido não está na lista de opções permitidas.'
            )
            dv.add(f'{col_letter}{START_ROW}:{col_letter}{END_ROW}')
            ws_data.add_data_validation(dv)
            
            # Ajusta a largura
            ws_data.column_dimensions[col_letter].width = 25
                
        # Ajustes de largura para outros campos (mantidos)
        elif verbose_name in ['Descrição', 'Observação editorial e núcleo', 'Observação da arte']:
            ws_data.column_dimensions[col_letter].width = 40
        elif 'Data' in verbose_name:
            ws_data.column_dimensions[col_letter].width = 18
        else:
            ws_data.column_dimensions[col_letter].width = 15

    # # 6. Salvar o arquivo
    # file_name = 'template_ilustracao_final.xlsx'
    # workbook.save(file_name)
    return workbook

def create_excel_with_data(queryset=None):
    """
    Gera um Excel com validação de dados (dropdowns) e, se fornecido, 
    preenche com os dados do queryset filtrado.
    """
    # 1. Obter e reordenar colunas conforme a lógica que já definiu
    colunas_originais = get_verbose_names(Ilustracao)
    COLUNAS_ILUSTRACAO = reordenar_colunas_para_excel(colunas_originais)

    if 'pk' not in [c.lower() for c in COLUNAS_ILUSTRACAO]:
        # Insere na segunda posição para respeitar a ordem do seu HTML
        COLUNAS_ILUSTRACAO.insert(1, 'pk')

    # 2. Configuração de dropdowns (mantendo a sua lógica original)
    field_options = get_choice_options(Ilustracao)
    dynamic_options = {}
    for verbose_name, value in field_options.items():
        if isinstance(value, str):
            dynamic_options[verbose_name] = get_foreign_key_values(value)

    # 3. Criar o Workbook e preparar a folha de opções oculta
    workbook = openpyxl.Workbook()
    ws_data = workbook.active
    ws_data.title = "Dados de Ilustracao"
    
    ws_options = workbook.create_sheet(title="Opcoes_Dropdown")
    ws_options.sheet_state = 'hidden'

    # Preencher ws_options e criar Defined Names para os dropdowns
    option_col = 1
    dropdown_references = {}
    all_dropdown_fields = {**field_options, **dynamic_options}

    for verbose_name, options_list in all_dropdown_fields.items():
        if isinstance(options_list, list) and options_list:
            ws_options.cell(row=1, column=option_col, value=verbose_name)
            for row_num, option in enumerate(options_list, 2):
                ws_options.cell(row=row_num, column=option_col, value=option)

            list_range = f'${get_column_letter(option_col)}$2:${get_column_letter(option_col)}${len(options_list) + 1}'
            dropdown_name = verbose_name.replace(' ', '_').replace('á', 'a').replace('é', 'e').replace('ç', 'c') + "_List"
            
            defined_name = DefinedName(name=dropdown_name, attr_text=f"'{ws_options.title}'!{list_range}")
            workbook.defined_names.add(defined_name)
            dropdown_references[verbose_name] = dropdown_name
            option_col += 1

    # 4. Escrever o Cabeçalho na folha principal
    ws_data.append(COLUNAS_ILUSTRACAO)
    ws_data.freeze_panes = 'A2'

    # 5. Preencher com dados do QuerySet
    if queryset:
        for objeto in queryset:
            linha = []
            for verbose_name in COLUNAS_ILUSTRACAO:
                # Caso especial para a coluna PK
                if verbose_name.lower() == 'pk':
                    valor = objeto.pk
                else:
                    # Tenta encontrar o campo pelo verbose_name
                    try:
                        campo = next(f for f in Ilustracao._meta.fields if f.verbose_name == verbose_name)
                        nome_campo = campo.name
                        
                        if hasattr(objeto, f'get_{nome_campo}_display'):
                            valor = getattr(objeto, f'get_{nome_campo}_display')()
                        else:
                            valor = getattr(objeto, nome_campo)
                    except StopIteration:
                        valor = ""

                # Tratamento para Foreign Keys e Objetos
                if hasattr(valor, 'pk') and valor:
                    valor = str(valor)
                
                # Tratamento para Datas (Formato Brasileiro)
                if hasattr(valor, 'strftime'):
                    valor = valor.replace(tzinfo=None) if valor else ""
                
                linha.append(valor if valor is not None else "")
            
            ws_data.append(linha)
            
            # Formatação de data nas células correspondentes
            row_idx = ws_data.max_row
            for col_idx, v_name in enumerate(COLUNAS_ILUSTRACAO, 1):
                if any(word in v_name for word in ['Data', 'Modificado', 'Criado']):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    cell.number_format = 'DD/MM/YYYY'

    # 6. Aplicar Validações de Dados (Dropdowns) e Ajustar Larguras
    START_ROW = 2
    END_ROW = 100  # Define até qual linha a validação funcionará

    for col_idx, verbose_name in enumerate(COLUNAS_ILUSTRACAO, 1):
        col_letter = get_column_letter(col_idx)
        
        # --- NOVO: APLICAÇÃO DA VALIDAÇÃO ---
        if verbose_name in dropdown_references:
            dv = DataValidation(
                type="list", 
                formula1=f'={dropdown_references[verbose_name]}', 
                allow_blank=True,
                showErrorMessage=True,
                errorTitle='Valor Inválido',
                error='O valor inserido não está na lista de opções permitidas.'
            )
            # Aplica a validação da linha inicial até a linha 1000 na coluna atual
            dv.add(f'{col_letter}{START_ROW}:{col_letter}{END_ROW}')
            ws_data.add_data_validation(dv)
            
            # Ajusta largura para colunas com dropdown
            ws_data.column_dimensions[col_letter].width = 25
        
        # --- AJUSTES DE LARGURA ADICIONAIS ---
        elif verbose_name.lower() == 'pk':
            ws_data.column_dimensions[col_letter].width = 8
        elif 'Data' in verbose_name or 'em' in verbose_name:
            ws_data.column_dimensions[col_letter].width = 18
        elif verbose_name in ['Descrição', 'Observação editorial e núcleo', 'Observação da arte']:
            ws_data.column_dimensions[col_letter].width = 40
        else:
            ws_data.column_dimensions[col_letter].width = 15

    return workbook
